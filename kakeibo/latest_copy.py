import glob
import os
import shutil
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd


def get_latest_row(ws, date_column=0):
    latest_row = 2
    for row in ws.iter_rows(min_row=latest_row + 1):
        if row[date_column].value is None:
            break
        latest_row += 1
    print(f'latest_row= {latest_row}')
    return latest_row


def get_latest_date(ws, latest_row, date_fmt, date_column=0):
    date = ws.cell(latest_row, date_column + 1).value
    if not isinstance(date, datetime):
        if 40000 < date < 100000:
            date = datetime(1899, 12, 30) + timedelta(days=date)
        else:
            date = datetime.strptime(str(date), date_fmt)
    print(f'latest_date= {date}')
    return date


def get_latest_descs(ws, latest_date, date_column, column_key_1, column_key_2):
    latest_descs = []
    for row in ws.iter_rows():
        if row[date_column].value == latest_date:
            latest_desc = str(row[column_key_1].value) + str(row[column_key_2].value)
            latest_descs.append(latest_desc)
    return latest_descs


# 明細が足りない場合の処理をおいおい追加したい


# depck = deplicate check 重複チェックのつもり
# num_fmtの取得が2個ある。これは2つの関数に切り分けられる気がする。
def update_data_depck(ws, date_column, column_key_1, column_key_2, latest_row, latest_date, latest_descs, csv_df, csv_date_fmt):
    num_fmt = []
    for cell in ws[latest_row:latest_row]:
        num_fmt.append(cell.number_format)
    print(f'num_fmt={num_fmt}')
    date_row = ws.cell(latest_row, date_column + 1).value

    for row in dataframe_to_rows(csv_df, index=None, header=None):
        if row[date_column] < latest_date:
            continue
        if (row[date_column] == latest_date) & (str(row[column_key_1] + str(row[column_key_2])) in latest_descs):
            continue
        latest_row += 1
        if isinstance(date_row, int):
            row[date_column] = int(row[date_column].strftime(csv_date_fmt))
        elif isinstance(date_row, str):
            row[date_column] = row[date_column].strftime(csv_date_fmt)
        for i, v in enumerate(row, 1):
            ws.cell(latest_row, i).value = v
            ws.cell(latest_row, i).number_format = num_fmt[i - 1]


def update_data(ws, date_column, column_key_1, column_key_2, latest_row, latest_date, latest_descs, csv_df, csv_date_fmt):
    num_fmt = []
    for cell in ws[latest_row:latest_row]:
        num_fmt.append(cell.number_format)
    print(f'num_fmt={num_fmt}')
    # date_row = ws.cell(latest_row, date_column + 1).value

    for row in dataframe_to_rows(csv_df, index=None, header=None):
        latest_row += 1
        if isinstance(row[date_column], int):
            row[date_column] = int(row[date_column].strftime(csv_date_fmt))
        elif isinstance(row[date_column], str):
            row[date_column] = row[date_column].strftime(csv_date_fmt)
        for i, v in enumerate(row, 1):
            ws.cell(latest_row, i).value = v
            ws.cell(latest_row, i).number_format = num_fmt[i - 1]

# 既知のバグ
# "日付"、"内容"、"残高"等のkeyで判定をしているため、
# 取り込んだcsvの最新の日付に、"内容"、"残高"が全く同じ行がある場合、正しく貼り付けできない


def latest_copy(wb, sheet_name, csv_file: str, csv_date_column_name, csv_date_fmt, date_column, key_column1, key_column2, csv_skiprows=0, csv_footerrows=0):
    ws = wb[sheet_name]
    latest_row = get_latest_row(ws)
    latest_date = get_latest_date(ws, latest_row, csv_date_fmt, date_column)
    latest_descs = get_latest_descs(ws, latest_date, date_column, key_column1, key_column2)
    df = pd.read_csv(csv_file, encoding='Shift-JIS', thousands=',', skiprows=csv_skiprows, skipfooter=csv_footerrows, engine='python')
    df[csv_date_column_name] = pd.to_datetime(df[csv_date_column_name], format=csv_date_fmt)
    update_data(ws, date_column, key_column1, key_column2, latest_row, latest_date, latest_descs, df, csv_date_fmt)


def latest_copy_depck(wb, sheet_name, csv_file: str, csv_date_column_name, csv_date_fmt, date_column, key_column1, key_column2, csv_skiprows=0, csv_footerrows=0):
    ws = wb[sheet_name]
    latest_row = get_latest_row(ws)
    latest_date = get_latest_date(ws, latest_row, csv_date_fmt, date_column)
    latest_descs = get_latest_descs(ws, latest_date, date_column, key_column1, key_column2)
    df = pd.read_csv(csv_file, encoding='Shift-JIS', thousands=',', skiprows=csv_skiprows, skipfooter=csv_footerrows, engine='python')
    df[csv_date_column_name] = pd.to_datetime(df[csv_date_column_name], format=csv_date_fmt)
    update_data_depck(ws, date_column, key_column1, key_column2, latest_row, latest_date, latest_descs, df, csv_date_fmt)


# if over_write:
#     wb.save(xlsx_name)
# else:
#     wb.save(f'new_{xlsx_name}')


def latest_csv_move_to(dir):
    home = os.getenv('HOME')
    download_folder = home + '/Downloads'
    list_of_files = glob.glob(download_folder + '/*.csv')
    latest_file = max(list_of_files, key=os.path.getctime)
    latest_file_name = os.path.basename(latest_file)
    new_path = shutil.move(latest_file, dir + '/' + latest_file_name)
    print('latest_file is moved')
    return new_path


if __name__ == '__main__':
    wb = load_workbook('hogehoge.xlsx')
    ws = wb['sheet_name']
    df = pd.read_csv('hogehoge.csv', encoding='Shift-JIS', thousands=',', skiprows=0)
    df['日付'] = pd.to_datetime(df['日付'], format='%Y/%m/%d')
    latest_row = get_latest_row(ws)
    latest_date = get_latest_date(ws, latest_row, 0)
    latest_descs = get_latest_descs(ws, latest_date, 0, 1, 4)
    update_data(ws, 0, 1, 4, latest_row, latest_date, latest_descs, df, '%Y/%m/%d')
    wb.save('new_hogehoge.xlsx')
