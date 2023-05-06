import detail_amazon_order_history
import latest_copy
from openpyxl import load_workbook


def execute(wb, sheet_name='Ama履歴', order_num_column=0, date_column=2, date_fmt='%Y-%m-%d', driver=None, skip_order_nums=[]):
    ws = wb[sheet_name]
    latest_row = latest_copy.get_latest_row(ws)
    print(f'latest_row={latest_row}')
    latest_order_num = ws.cell(latest_row, order_num_column + 1).value
    print(f'latest_order_num={latest_order_num}')
    detail_amazon_order_history.fetch(latest_order_num, driver, skip_order_nums)
    df = detail_amazon_order_history.history_to_df()
    latest_copy.update_data(ws, date_column, latest_row, df, date_fmt, index=False)
    return driver


if __name__ == '__main__':
    wb = load_workbook("./data/ama_his_test.xlsx")
    skip_order_nums = ['']
    execute(wb, 'Ama履歴', skip_order_nums=skip_order_nums)
    wb.save("./data/ama_his_test_update.xlsx")

    print('done')
